import openpyxl
import json

CITIES = [
    "تهران",
    "اصفهان",
    "مازندران",
    "قم",
    "گیلان",
    "البرز",
    "مرکزی",
    "خراسان رضوی",
    "سمنان",
    "آذربایجان  شرقی",
    "قزوین",
    "یزد",
    "فارس",
    "لرستان",
    "خوزستان",
    "گلستان",
    "آذربایجان غربی",
    "زنجان",
    "اردبیل",
    "کردستان",
    "همدان",
    "کرمانشاه",
    "خراسان جنوبی",
    "کرمان",
    "هرمزگان",
    "ایلام",
    "خراسان شمالی",
    "سیستان و بلوچستان",
    "چهارمحال و بختیاری",
    "بوشهر",
    "کهگیلویه و بویراحمد"
]

# Parameters
PATIENTS_ROW = 32
DEAD_ROW = 33
RECOVERED_ROW = 34
file_name = 'data.xlsx'
MAX_COLUMN_COUNT = 10000

# Open excel file
book = openpyxl.load_workbook(file_name)
sheet = book.active

# Get proper column
CurrentColumnIndex = -1
for row in sheet.iter_rows(min_row=1, min_col=1, max_row=1, max_col=MAX_COLUMN_COUNT):
    for cell in row:
        if cell.value == None:
            CurrentColumnIndex = cell.column
            break

# Get Patients, Dead and Recovered and write them to sheet
patients = input("patients: ")
dead = input("dead: ")
recovered = input("recovered: ")
sheet.cell(row=PATIENTS_ROW, column=CurrentColumnIndex).value = int(patients)
sheet.cell(row=DEAD_ROW, column=CurrentColumnIndex).value = int(dead)
sheet.cell(row=RECOVERED_ROW, column=CurrentColumnIndex).value = int(recovered)

# Get news text
print("Enter News Text (Enter \"-1\" to end):")
NEWS_TEXT = []
while True:
    inp = input()
    if inp == "-1":
        break
    NEWS_TEXT.append(inp)

# Check if each city is not in a separate line. if so, split each city in separate line
if len(NEWS_TEXT) < 10:
    tmp = NEWS_TEXT.copy()
    NEWS_TEXT.clear()
    for item in tmp:
        NEWS_TEXT += item.split("،") # Farsi Comma is default


# Mining news text
ignored_city = CITIES.copy() # list of unfounded cities
for city in CITIES:
    city_nospace = "".join(city.split())  # delete all white spaces
    for news in NEWS_TEXT:
        news_nospace = "".join(news.split())  # delete all white spaces
        if news_nospace.__contains__(city_nospace):
            if city == "کرمان" and news.__contains__("کرمانشاه"):  # This is a special case, should check manually
                continue
            nwsns_sp = news_nospace.split(":")
            for split_part in nwsns_sp:
                try:
                    int(split_part)  # find the number

                    for i in range(1, PATIENTS_ROW): # Search for city row
                        if sheet.cell(row=i, column=1).value == city: #column 1 is name
                            sheet.cell(row=i, column=CurrentColumnIndex).value = sheet.cell(row=i,
                                                                                            column=CurrentColumnIndex-1).value + int(split_part)
                            ignored_city.remove(city)
                            break
                except ValueError:
                    continue # Throw int(split_part)
            break

# Print unfounded cities
if len(ignored_city) > 0:
    print("Can't find below cities:")
    print("\t", ignored_city)
    for ign_city in ignored_city: # Fill cell with previous cell value
        for i in range(1, PATIENTS_ROW):
            if sheet.cell(row=i, column=1).value == ign_city: #column 1 is name
                sheet.cell(row=i, column=CurrentColumnIndex).value = sheet.cell(row=i, column=CurrentColumnIndex-1).value
                break


# Make json file
data = {}
data['provinces'] = []

for row in sheet.iter_rows(min_row=1, min_col=2, max_row=PATIENTS_ROW-1, max_col=CurrentColumnIndex):
    pat = []
    for cell in row:
        pat.append(cell.value)

    data['provinces'].append({
        'name': sheet.cell(row=row[0].row, column=1).value,
        'patients': pat
    })


All = [] # 0 is patiens, 1 is dead, 2 is recovered
for row in sheet.iter_rows(min_row=PATIENTS_ROW, min_col=2, max_row=RECOVERED_ROW, max_col=CurrentColumnIndex):
    temp = []
    for cell in row:
        temp.append(cell.value)
    All.append(temp)

data['iran'] = {
    'patients': All[0],
    'dead': All[1],
    'recovered': All[2]
}

with open('infected.json', 'w', encoding='utf8') as outfile:
    json.dump(data, outfile, ensure_ascii=False)

# Save excel file
print("Completed!")
book.save(file_name)